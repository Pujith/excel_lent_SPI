import openpyxl

def open_workbook():
    workbook = openpyxl.load_workbook('file.xlsx')
    file_sheet = workbook.active
    for row in range(2,86):
        cell_value = file_sheet.cell(row, 1)
        totalCount = 0
        for column in range(2,909):
            cell_color = file_sheet.cell(column=column, row=row)
            fgColor = cell_color.fill.fgColor.index
            if fgColor == 'FF35A654' or fgColor == 'FF6AA84F' or fgColor == 'FF34A853' or fgColor == 'FF34A754' or fgColor == 'FF71AD47' or fgColor == '':
                totalCount = totalCount+1
        print(f'{cell_value.value}, {totalCount}')


if __name__ == '__main__':
    open_workbook()
